from tkinter import *
from tkinter import messagebox
import nepali_datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,numbers
import random,os
from tkinter import END
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from win32com import client
from ttkbootstrap import Style
import ttkbootstrap as ttk
import amounttowords
import db

if not os.path.exists('billno.txt'):
    with open("billno.txt","w+") as f:
        f.write('1')
else:
    with open("billno.txt","r+") as f:
        billno =  f.read()
        billno = int(billno)
        with open("billno.txt","w") as fl:
            updatebill = billno+1
            update_bill = fl.write(str(updatebill))

today = str(nepali_datetime.date.today())

if not os.path.exists('bills'):
    os.mkdir('bills')

if not os.path.exists('bills'):
    os.mkdir('PDFbills')

wb = load_workbook('brocode.xlsx')
ws = wb.active 

def make_pdf():
    try:
        global billno
        app = client.DispatchEx("Excel.Application")
        app.Interactive = False
        app.Visible = False
        app.DisplayAlerts = False  # 🚫 Prevents prompts like "Do you want to save?"

        file = r"C:\Users\archlinux\Desktop\Mount-Care\brocode.xlsx"
        excel_dir = os.path.dirname(file)

        billno = str(billno)
        billdir = "PDFbills/"
        billno += ".pdf"
        billdir += billno

        pdf_file_path = os.path.join(excel_dir, billdir)

        if not os.path.exists(os.path.dirname(pdf_file_path)):
            os.makedirs(os.path.dirname(pdf_file_path))

        workbook = app.Workbooks.Open(file)
        sheet = workbook.ActiveSheet

        #  Fit to one page without zooming out too much
        sheet.PageSetup.Zoom = False
        sheet.PageSetup.FitToPagesWide = 1
        sheet.PageSetup.FitToPagesTall = 1

        # Center content horizontally

        sheet.ExportAsFixedFormat(0, pdf_file_path)

        workbook.Close(SaveChanges=False)  # 🔐 Don't save changes

        print("Successfully converted to PDF!")

    except Exception as e:
        print("Error:", e)



def send_email():
    pass

def print_bill():
    pass

#Search Function
def search_bill():
    for i in os.listdir('bills/'):
        if i.split('.')[0] == BillEntry.get():
            f = open(f'bills/{i}','r')
            textarea.delete(1.0,END)
            for data in f:
                textarea.insert(END,data)
            f.close()
            break
    else:
        messagebox.showerror("Error",'Invalid Bill Number')

def reverse_bill():
    st = start
    ws.delete_rows(st,mainbillcnt)

    ofst = 22
    ws.delete_rows(ofst,offerbillcnt)


def write_to_excel():
    # Main Bill
    temp = []
    Otemp = []

    global start,end,Ostart,Oend
    start = 13
    end = 13+mainbillcnt

    total = end+1
    discount = end+2
    amtpay = end+3

    for product, qty in mainbill_qty.items():
        temp.append(product)
    for product, qty in offer_qty.items():
        Otemp.append(product)

    
    top = Side(border_style='thin')
    buttom = Side(border_style='thin')
    right = Side(border_style='thin')
    left = Side(border_style='thin')

    border = Border(top=top,right=right,bottom=buttom,left=left)

    alignment = Alignment(horizontal='center',shrink_to_fit=True)
    shrink = Alignment(shrink_to_fit=True)
        
    i = 0
    row = start
    while row < end+1 and i<len(temp):
        ws.insert_rows(row)

        ws['C'+str(row)] = i+1
        ws['D'+str(row)] = temp[i]
        ws['E'+str(row)] = mainbill_qty[temp[i]]
        ws['F'+str(row)] = all_product[temp[i]][0]
        ws['G'+str(row)] = mainbill_tprice[temp[i]]

        ws['E'+str(row)].alignment = alignment
        ws['F'+str(row)].alignment = shrink
        ws['G'+str(row)].alignment = shrink

        ws['C'+str(row)].font = Font(bold=True,size=14)
        ws['D'+str(row)].font = Font(bold=True,size=14)
        ws['E'+str(row)].font = Font(bold=True,size=14)
        ws['F'+str(row)].font = Font(bold=True,size=14)
        ws['G'+str(row)].font = Font(bold=True,size=14)

        ws['B'+str(row)].fill = PatternFill(fill_type= 'solid',start_color='9c2c95')
        ws['C'+str(row)].fill = PatternFill(fill_type= 'solid',start_color='9c2c95')
        ws['D'+str(row)].fill = PatternFill(fill_type= 'solid',start_color='9c2c95')
        ws['E'+str(row)].fill = PatternFill(fill_type= 'solid',start_color='9c2c95')
        ws['F'+str(row)].fill = PatternFill(fill_type= 'solid',start_color='9c2c95')
        ws['G'+str(row)].fill = PatternFill(fill_type= 'solid',start_color='9c2c95')
        ws['H'+str(row)].fill = PatternFill(fill_type= 'solid',start_color='9c2c95')

        ws['C'+str(row)].border = border
        ws['D'+str(row)].border = border
        ws['E'+str(row)].border = border
        ws['F'+str(row)].border = border
        ws['G'+str(row)].border = border


        i+=1
        row+=1

    ws['G'+str(total)] = Totals
    ws['G'+str(discount)] = Discounts
    ws['G'+str(amtpay)] = Amountpays
    words = amounttowords.number_to_words(int(float(Amountpay)))

    if len(words)>=44:
        temp = words
        ans = ""
        lsts = words.split() #"sahil is a bad boy" "sahil is"
        while len(temp)>=44:
            ans = lsts.pop() + " " + ans
            temp = " ".join(lsts)
        ws['E'+str(amtpay+1)] = temp
        ws['E'+str(amtpay+2)] = ans
    else:
        ws['E'+str(amtpay+1)] = words

    total = end
    discount = end
    amtpay = end+1

def save_bill():
    result = messagebox.askyesno('Confirm','Do you want to save the bill?')
    if result:
        bill_content = textarea.get(1.0,END)
        file = open(f'bills/{billno}.txt','w')
        file.write(bill_content)
        file.close()
        messagebox.showinfo(f'Sucess,{billno} is saved sucessfully')
        wb.save('brocode.xlsx')

#writing in excel file
ws['E7'].value = billno

def bill_area():
    if nameEntry.get()=='' or MoidEntry.get() == '' or dealerEntry.get() == '' or sdealerEntry.get() == '' or AddressEntry.get() == '':
        messagebox.showerror('Error','Customer Detail Are Required!')
    elif TotalEntry.get() == '':
        messagebox.showerror('No Product Are Selected!')
    elif TotalEntry.get() == '0':
        messagebox.showerror('No Product Are Selected!')
    elif TotalEntry1.get() == '' or TotalEntry1.get() =='0.0':
        messagebox.showerror('No Offer Product Are Selected!')
    else:
        textarea.delete(1.0,END)
        textarea.insert(END,'\t*** Mount Care International Pvt Ltd ***\n')
        textarea.insert(END,'\t\tKathmandu-3, Maharajung Chok \n')
        textarea.insert(END,'\t    Phone Number : 9851177355,9823681553 \n')
        textarea.insert(END,f'\n  Bill Number :\t {billno:>17}')
        textarea.insert(END,f'\n  Customer Name :\t {nameEntry.get():>15}')
        textarea.insert(END,f'\n  Customer Id :\t {MoidEntry.get():>17}')
        textarea.insert(END,f'\n  Customer Address :\t {AddressEntry.get():>12}\n')
        textarea.insert(END,f'\n  ====================================================== ')
        textarea.insert(END,f'\n  Product    \t\t\tQuantity                \t\t\tPrice')
        textarea.insert(END,f'\n  ====================================================== ')

        ws['E9'].value = MoidEntry.get() #Customer ID
        ws['E8'].value = nameEntry.get() #Customer Name
        ws['E10'].value = AddressEntry.get() #Customer Address
        ws['G7'].value = today # Date
        
        global mainbillcnt,offerbillcnt,mainbill_qty,mainbill_tprice
        mainbill_tprice = {} # Store the Price of the Product Inserted by Input Field
        mainbill_qty = {} # Store the Quantity of the product
        mainbillcnt = 0 # Entry in product in Main
        offerbillcnt = 0

        for pro,item in main_product.items():
            if item.get() != '0' and item.get() != '':
                mainbillcnt+=1
                textarea.insert(END,f'\n {pro}\t\t   {item.get():>13}  \t\t\t{all_product[pro][0]:>25}')
                mainbill_qty[pro] = item.get()
                mainbill_tprice[pro] = all_product[pro][0]

        textarea.insert(END,f'\n  ------------------------------------------------------')

        global Totals,Offers,Discounts,Amountpays,Amountpay

        Totals = TotalEntry.get()
        Offers = DisEntry.get()
        Discounts = DiscountEntry.get()
        Amountpay = AmountpayableEntry.get()

        Totals = "Rs " + Totals
        Offers = "Rs " + Offers
        Discounts = "Rs " + Discounts
        Amountpays = "Rs "+ Amountpay

        textarea.insert(END,"\n")
        textarea.insert(END,f'\n  Total Amount  \t\t     {Totals:>24}')
        textarea.insert(END,f'\n  Offer  \t\t     {Offers:>25}')
        textarea.insert(END,f'\n  Discount  \t\t     {Discounts:>25}\n')
        textarea.insert(END,f'\n  Amount Payable  \t\t     {Amountpays:>22}')

        textarea.insert(END,'\n\t\t### Offer Products ###\n')
        textarea.insert(END,f'\n  ====================================================== ')
        textarea.insert(END,f'\n  Product    \t\t\tQuantity\n')
        textarea.insert(END,f'\n  ====================================================== ')

        for product,qty in offer_qty.items():
            if qty>0:
                offerbillcnt+=1
                textarea.insert(END, f'\n  {product}  {qty:>20}')

        write_to_excel()
        save_bill()
        make_pdf()
        ws['E'+str(end+5)] = '=""'
        reverse_bill()
        wb.save('brocode.xlsx')

#GUI part
root=Tk()
root.title('Mount Care System')
root.geometry('1920x1080')
root.iconbitmap('icon.ico')
style = Style(theme='darkly')
headingLabel=Label(root,text='Mount Care Billing System',font=('times new roman',30,'bold'),bg='gray20',fg='chartreuse3',bd=13,relief=GROOVE)
headingLabel.pack(fill=X)

customer_details_Frame=LabelFrame(root,text='Customer Details',font=('times new roman',18,'bold'),fg='chartreuse3',border=8,relief=GROOVE,bg='gray20')
customer_details_Frame.pack(fill=X)

nameLabel =Label(customer_details_Frame,text='Name',font=('times new roman',15,'bold'),bg='gray20',fg='white')
nameLabel.grid(row=0,column=0)


nameEntry = Entry(customer_details_Frame,font=('arial',15),bd=7,width=18) #name field
nameEntry.grid(row=0,column=1,padx=(0,20))


AddressLabel =Label(customer_details_Frame,text='Address',font=('times new roman',15,'bold'),bg='gray20',fg='white')
AddressLabel.grid(row=0,column=2,pady=2)

AddressEntry = Entry(customer_details_Frame,font=('arial',15),bd=7,width=10) #name field
AddressEntry.grid(row=0,column=3,padx=(0,20))

MoidLabel =Label(customer_details_Frame,text='Mo Id',font=('times new roman',15,'bold'),bg='gray20',fg='white')
MoidLabel.grid(row=0,column=4,pady=2)

MoidEntry = Entry(customer_details_Frame,font=('arial',15),bd=7,width=18) #name field
MoidEntry.grid(row=0,column=5,padx=(0,20))

dealerLabel =Label(customer_details_Frame,text='Dealer',font=('times new roman',15,'bold'),bg='gray20',fg='white')
dealerLabel.grid(row=0,column=6,pady=2)

dealerEntry = Entry(customer_details_Frame,font=('arial',15),bd=7,width=18) #name field
dealerEntry.grid(row=0,column=7,padx=(0,20))

sdealerLabel =Label(customer_details_Frame,text='SDealer',font=('times new roman',15,'bold'),bg='gray20',fg='white')
sdealerLabel.grid(row=0,column=8,pady=2)

sdealerEntry = Entry(customer_details_Frame,font=('arial',15),bd=7,width=18) #name field
sdealerEntry.grid(row=0,column=9,padx=(0,20))

BillLabel =Label(customer_details_Frame,text='Bill Number',font=('times new roman',15,'bold'),bg='gray20',fg='white')
BillLabel.grid(row=0,column=10,pady=2)

BillEntry = Entry(customer_details_Frame,font=('arial',15),bd=7,width=18) #name field
BillEntry.grid(row=0,column=11,padx=(0,20))

searchButton = Button(customer_details_Frame,text='SEARCH',font=("arial",13,'bold'),bd=7,width=10,command=search_bill)
searchButton.grid(row=0,column=13,padx=(0,13),pady=8)


productFrame=Frame(root)
productFrame.pack()

FoodSupplementFrame = LabelFrame(productFrame,text='Food Supplement',font=('times new roman',18,'bold'),fg='chartreuse3',border=8,relief=GROOVE,bg='gray20')
FoodSupplementFrame.grid(row=0,column=0)


all_product = db.dbconnect()



CosmeticFrame = LabelFrame(productFrame,text='Cosmetic',font=('times new roman',18,'bold'),fg='chartreuse3',border=8,relief=GROOVE,bg='gray20')
CosmeticFrame.grid(row=0,column=1,sticky='N')


BeverageFrame = LabelFrame(productFrame,text='Beverage',font=('times new roman',18,'bold'),fg='chartreuse3',border=8,relief=GROOVE,bg='gray20')
BeverageFrame.grid(row=0,column=2,sticky='N')

global tatalsum
def sum_total(event, main_product, discount_entry):
    total = 0
    for pro, entry in main_product.items():
        try:
            value = entry.get().strip()
            total += all_product[pro][0] * int(value) if value else 0
        except ValueError:
            continue  # skip if invalid input

    # Update TotalEntry
    AmountpayableEntry.delete(0, END)
    TotalEntry.delete(0, END)
    DiscountEntry.delete(0,END)

    # Apply discount
    dispercent = discount_entry.get().strip()
    try:
        discount = float(dispercent)
    except ValueError:
        discount = 0  # If not a number, treat as 0%

    discounted_total = total - (discount / 100) * total
    dis = discount/100*total


    Amountpay = '{:.1f}'.format(discounted_total)
    dis = '{:.1f}'.format(dis)
    total = '{:.1f}'.format(total)

    DiscountEntry.insert(0,dis)
    TotalEntry.insert(0,total)
    AmountpayableEntry.insert(0, Amountpay)

billFrame=Frame(productFrame,bd=8,relief=GROOVE)
billFrame.grid(row=0,column=3,padx=10)

billareaLabel=Label(billFrame,text='Bill Area',font=('times new roman',18,'bold'),bd=7,relief=GROOVE)
billareaLabel.pack(fill=X)

scrollbar=Scrollbar(billFrame,orient=VERTICAL)
scrollbar.pack(side=RIGHT,fill=Y)

textarea=Text(billFrame,height=35,width=60,yscrollcommand=scrollbar.set) #28,55
textarea.pack()
scrollbar.config(command=textarea.yview)


billFrame=Frame(productFrame,bd=8,relief=GROOVE)
billFrame.grid(row=0,column=4,padx=10)

OfferFrame = LabelFrame(productFrame, text='Offer Section', font=('times new roman', 18, 'bold'), fg='chartreuse3',
                         border=8, relief=GROOVE, bg='gray20')
OfferFrame.grid(row=0, column=4)

# Creating a canvas and scrollbar for the Offer Section
offer_canvas = Canvas(OfferFrame, bg='gray20', bd=0, relief=GROOVE, width=400, height=580)
offer_scrollbar = Scrollbar(OfferFrame, orient=VERTICAL, command=offer_canvas.yview)
offer_scrollbar.pack(side=RIGHT, fill=Y)
offer_canvas.pack(side=LEFT, fill=BOTH, expand=True)
offer_canvas.configure(yscrollcommand=offer_scrollbar.set)

# Create a frame inside the canvas to hold the widgets
offer_frame = Frame(offer_canvas, bg='gray20')
offer_canvas.create_window((0, 0), window=offer_frame, anchor='nw')

def create_product_entries(frame, products):
    entry_dict = {}
    i = 0
    for pro in products:
        label = Label(frame, text=pro, font=('times new roman', 15, 'bold'), bg='gray20', fg='white')
        label.grid(row=i, column=0, sticky='w', pady=5, padx=10)

        entry = Entry(frame, font=('times new roman', 15, 'bold'), bd=5, width=10)
        entry.grid(row=i, column=1, pady=5)
        entry.insert(0,0)
        entry_dict[pro] = entry  # Associate each entry field with its product name
        i+=1

    for entry in entry_dict.values():
        entry.bind('<FocusOut>', lambda event: calculate_total(entry_dict))  # Bind FocusOut event

def calculate_total(entry_dict):
    total = 0
    global offer_qty
    offer_qty = {}
    for product, entry in entry_dict.items():
        value = int(entry.get()) * all_product[product][0]
        if value != 0 and product not in offer_qty: offer_qty[product] = int(entry.get())
        total+=value

    TotalEntry1.delete(0,END)  # You can replace print with any action you want to take with the total value
    TotalEntry1.insert(0,f"{total}")
offer_product =  create_product_entries(offer_frame, all_product)

# Configuring canvas scrolling region
offer_frame.update_idletasks()
offer_canvas.config(scrollregion=offer_canvas.bbox("all"))

# Binding mouse wheel event to canvas
def on_mousewheel(event):
    offer_canvas.yview_scroll(int(-1 * (event.delta / 130)), "units")

offer_canvas.bind_all("<MouseWheel>", on_mousewheel)

BillmenuFrame = LabelFrame(root,text='Bill Menu',font=('times new roman',18,'bold'),fg='chartreuse3',border=8,relief=GROOVE,bg='gray20')
BillmenuFrame.pack(pady=10,side=TOP, anchor=NW)

TotalLabel=Label(BillmenuFrame,text='Total',font=('times new roman',14,'bold'),bg='gray20',fg='white')
TotalLabel.grid(row=0,column=0,pady=5,padx=10,sticky='w')

TotalEntry=Entry(BillmenuFrame,font=('times new roman',14,'bold'),width=15,bd=5)
TotalEntry.grid(row=0,column=1,pady=5,padx=10)

DiscountLabel=Label(BillmenuFrame,text='Discount Amt',font=('times new roman',14,'bold'),bg='gray20',fg='white')
DiscountLabel.grid(row=2,column=0,pady=5,padx=10,sticky='w')

DiscountEntry=Entry(BillmenuFrame,font=('times new roman',14,'bold'),width=15,bd=5)
DiscountEntry.grid(row=2,column=1,pady=5,padx=10)

AmountpayableLabel=Label(BillmenuFrame,text='Amount Payable',font=('times new roman',14,'bold'),bg='gray20',fg='white')
AmountpayableLabel.grid(row=3,column=0,pady=5,padx=10,sticky='w')

AmountpayableEntry=Entry(BillmenuFrame,font=('times new roman',14,'bold'),width=15,bd=5)
AmountpayableEntry.grid(row=3,column=1,pady=5,padx=10)

#begin

dislabel=Label(BillmenuFrame,text='Discount %',font=('times new roman',14,'bold'),bg='gray20',fg='white')
dislabel.grid(row=0,column=2,pady=5,padx=10,sticky='w')

global DisEntry
DisEntry=Entry(BillmenuFrame,font=('times new roman',14,'bold'),width=15,bd=5)
DisEntry.grid(row=0,column=3,pady=5,padx=10)
DisEntry.insert(5,5)


buttonFrame=Frame(BillmenuFrame,bd=8,relief=GROOVE)
buttonFrame.grid(row=0,column=4,rowspan=7)

totalButton=Button(buttonFrame,text='TOTAL',font=('arial',18,'bold'),bg='chartreuse3',fg='white',bd=5,width=8)
totalButton.grid(row=0,column=0,pady=30,padx=5)

#here uis t
# Here is the existing code for the first bill menu...
lev = Label(BillmenuFrame, text='Offer Bill', font=('times new roman', 18, 'bold'), bg='gray20', fg='chartreuse3')
lev.grid(row=0, column=5, pady=5, padx=(147,0))

TotalLabel2 = Label(BillmenuFrame, text='Offer Total', font=('times new roman', 13, 'bold'), bg='gray20', fg='white')
TotalLabel2.grid(row=2, column=5, pady=5, padx=(150,0))

TotalEntry1=Entry(BillmenuFrame,font=('times new roman',14,'bold'),width=10,bd=5)
TotalEntry1.grid(row=2,column=6,pady=5,padx=(50,1000))

ClearButton2=Button(BillmenuFrame,text='CLEAR',font=('arial',14,'bold'),bg='chartreuse3',fg='white',bd=5,width=8)
ClearButton2.grid(row=3,column=6,pady=3,padx=(50,1000))

i = 0
global main_product
main_product = {}

for pro in all_product:
    category = all_product[pro][-1]

    if category == 'food supplement':
        frame = FoodSupplementFrame
    elif category == 'Cosmetic':
        frame = CosmeticFrame
    elif category == 'Beverage':
        frame = BeverageFrame
    else:
        continue

    label = Label(frame, text=pro, font=('times new roman', 15, 'bold'), bg='gray20', fg='white')
    label.grid(row=i, column=0, pady=9, padx=10, sticky='w')

    entry = Entry(frame, font=('times new roman', 15, 'bold'), width=10, bd=5)
    entry.grid(row=i, column=1, pady=9, padx=10)
    entry.insert(0, 0)

    #bind key event and pass main_product
    entry.bind("<KeyRelease>", lambda event, mp=main_product, dg=DisEntry: sum_total(event, mp, dg))

    main_product[pro] = entry
    i += 1
entry.bind("<KeyRelease>", lambda event, mp=main_product, dg=DisEntry: sum_total(event, mp, dg))
# Bind the discount entry
DisEntry.bind("<KeyRelease>", lambda event, mp=main_product, dg=DisEntry: sum_total(event, mp, dg))

BillButton=Button(buttonFrame,text='BILL',font=('arial',18,'bold'),bg='chartreuse3',fg='white',bd=5,width=8,command=bill_area)
BillButton.grid(row=0,column=1,pady=30,padx=5)

EmailButton=Button(buttonFrame,text='EMAIL',font=('arial',18,'bold'),bg='chartreuse3',fg='white',bd=5,width=8)
EmailButton.grid(row=0,column=2,pady=30,padx=5)

PrintButton=Button(buttonFrame,text='PRINT',font=('arial',18,'bold'),bg='chartreuse3',fg='white',bd=5,width=8,command=make_pdf)
PrintButton.grid(row=0,column=3,pady=30,padx=10,)

ClearButton=Button(buttonFrame,text='CLEAR',font=('arial',18,'bold'),bg='chartreuse3',fg='white',bd=5,width=8)
ClearButton.grid(row=0,column=4,pady=30,padx=5)
root.mainloop()